"""读取 Excel 并生成 Drama 列表。"""

import re
from pathlib import Path
from typing import List

from openpyxl import load_workbook

from models import Drama


# 表头 → 字段名 的映射
HEADER_MAP = {
    "原剧名": "original_title",
    "原简介": "original_desc",
    "作者": "author",
    "分类": "category",
    "原海报URL": "original_cover_url",
    "翻译语言": "translation_language",
    "翻译后剧名": "title",
    "翻译后简介": "description",
    "新海报URL": "cover_url",
    "剧集数": "episode_count",
    "剧集 URL": "video_urls",
    "剧集URL列表": "video_urls",
}


def _normalize_header(value) -> str:
    """把表头统一成字符串并去除首尾空白。"""
    if value is None:
        return ""
    return str(value).strip()


def _split_video_urls(raw) -> List[str]:
    """把单元格里的视频 URL 按换行/空格/逗号分割，过滤空值。"""
    if raw is None:
        return []
    text = str(raw)
    # 支持换行、回车、逗号、中文逗号、空格分隔
    urls = re.split(r"[\r\n,，\s]+", text)
    return [u.strip() for u in urls if u.strip().startswith("http")]


def read_excel(path: Path) -> List[Drama]:
    """读取 Excel，返回 Drama 列表（跳过表头行）。"""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # 读取表头
    headers = [_normalize_header(cell.value) for cell in ws[1]]
    col_map = {}
    for idx, h in enumerate(headers):
        if h in HEADER_MAP:
            col_map[HEADER_MAP[h]] = idx

    missing = set(HEADER_MAP.values()) - set(col_map.keys())
    if missing:
        raise ValueError(f"Excel 缺少必要列: {missing}，实际表头: {headers}")

    dramas: List[Drama] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 如果整行第一个关键列为空，认为已到末尾
        title_val = row[col_map["title"]] if "title" in col_map else None
        if title_val is None or str(title_val).strip() == "":
            continue

        episode_count_raw = row[col_map["episode_count"]]
        try:
            episode_count = int(float(str(episode_count_raw).strip()))
        except (ValueError, TypeError):
            episode_count = 0

        drama = Drama(
            row_idx=row_idx,
            original_title=str(row[col_map["original_title"]] or "").strip(),
            original_desc=str(row[col_map["original_desc"]] or "").strip(),
            author=str(row[col_map["author"]] or "").strip(),
            category=str(row[col_map["category"]] or "").strip(),
            original_cover_url=str(row[col_map["original_cover_url"]] or "").strip(),
            translation_language=str(row[col_map["translation_language"]] or "").strip(),
            title=str(row[col_map["title"]] or "").strip(),
            description=str(row[col_map["description"]] or "").strip(),
            cover_url=str(row[col_map["cover_url"]] or "").strip(),
            episode_count=episode_count,
            video_urls=_split_video_urls(row[col_map["video_urls"]]),
        )
        dramas.append(drama)

    return dramas


if __name__ == "__main__":
    import json
    from pathlib import Path

    dramas = read_excel(Path("export_20260705_180246.xlsx"))
    print(f"共读取 {len(dramas)} 部剧")
    for d in dramas:
        print(f"[{d.row_idx}] {d.title} | 集数:{d.episode_count} | 视频:{len(d.video_urls)}个")
